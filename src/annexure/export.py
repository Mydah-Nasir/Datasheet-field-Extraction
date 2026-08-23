"""Export logic for AnnexureRecords into Excel, CSV, and JSON formats."""

import csv
import io
import json
from collections.abc import Sequence

import openpyxl
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.worksheet.worksheet import Worksheet

from src.annexure.models import AnnexureRecord


def export_to_json(records: Sequence[AnnexureRecord]) -> bytes:
    """Export AnnexureRecords to a JSON byte string."""
    data = [record.model_dump() for record in records]
    return json.dumps(data, indent=2).encode("utf-8")


def export_to_csv(records: Sequence[AnnexureRecord]) -> bytes:
    """Export AnnexureRecords to a CSV byte string."""
    if not records:
        return b""

    output = io.StringIO()
    # Use the model fields as the header
    field_names = list(AnnexureRecord.model_fields.keys())
    # You might want to map these to the actual headers from field descriptions
    headers = [AnnexureRecord.model_fields[name].description or name for name in field_names]

    writer = csv.writer(output)
    writer.writerow(headers)

    for record in records:
        row = [getattr(record, name) for name in field_names]
        writer.writerow(row)

    return output.getvalue().encode("utf-8")


def _apply_excel_formatting(ws: Worksheet) -> None:
    """Apply engineering styling to the Annexure worksheet."""
    # Bold header font
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Style the title row (row 1)
    title_cell = ws.cell(row=1, column=1)
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="left", vertical="center")

    # Style headers (row 2)
    for cell in ws[2]:
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    # Freeze panes below the header
    ws.freeze_panes = "A3"

    # Set filters
    ws.auto_filter.ref = f"A2:{openpyxl.utils.get_column_letter(ws.max_column)}{ws.max_row}"

    # Auto-adjust column widths based on headers and data (approximate)
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            # Skip the title row for width calculation
            if cell.row == 1:
                continue
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        adjusted_width = max_length + 2
        # Cap width
        if adjusted_width > 50:
            adjusted_width = 50
        ws.column_dimensions[col_letter].width = adjusted_width


def export_to_excel(records: Sequence[AnnexureRecord]) -> bytes:
    """Export AnnexureRecords to an Excel (.xlsx) byte string."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Annexure"

    # Add title row
    ws.append(["BID PROJECT – MECHANICAL DATASHEET ANNEXURE"])

    if records:
        # Add headers
        field_names = list(AnnexureRecord.model_fields.keys())
        headers = [AnnexureRecord.model_fields[name].description or name for name in field_names]
        ws.append(headers)

        # Add data rows
        for record in records:
            row = [getattr(record, name) for name in field_names]
            ws.append(row)

        # Apply borders to data rows
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = thin_border

                # If numeric, ensure it's saved as a number in Excel, not string
                if isinstance(cell.value, (int, float)):
                    cell.data_type = "n"

    # Format worksheet
    _apply_excel_formatting(ws)

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
