"""Unit tests for annexure export generation."""

import csv
import io
import json

import openpyxl

from src.annexure.export import export_to_csv, export_to_excel, export_to_json
from src.annexure.models import AnnexureRecord


def create_mock_records():
    """Create a list of mock AnnexureRecords for testing."""
    return [
        AnnexureRecord(
            tag_no="V-101",
            description="Separator",
            ref_data_sheet="DS-001",
            design_code="ASME Sec VIII",
            moc="SA-516 Gr 70",
            qty=2,
            orientation="VERTICAL",
            vessel_id_mm=1500.0,
            vessel_tl_tl_length_mm=4000.0,
            shell_min_thk_mm=20.0,
            head_min_thk_mm=22.0,
            head_type="2:1 Elliptical",
            nozzle_type="Flanged",
            impact_tested="YES",
            rt="FULL",
            pwht="YES",
            support_type="Skirt",
            painting_external="System 1",
            painting_internal="System 2",
            weight_tons_each=15.5,
        ),
        AnnexureRecord(
            tag_no="V-102",
            description="Storage Tank",
            ref_data_sheet="DS-002",
            design_code="API 650",
            moc="A36",
            qty=1,
            orientation="HORIZONTAL",
            vessel_id_mm=3000.0,
            vessel_tl_tl_length_mm=6000.0,
            shell_min_thk_mm=10.0,
            head_min_thk_mm=12.0,
            head_type="Flat",
            nozzle_type="Threaded",
            impact_tested="NO",
            rt="NONE",
            pwht="NO",
            support_type="Saddle",
            painting_external="System 3",
            painting_internal="None",
            weight_tons_each=5.0,
        ),
    ]


def test_export_to_json():
    """Test exporting to JSON format."""
    records = create_mock_records()
    json_bytes = export_to_json(records)

    # Verify we can decode it
    json_str = json_bytes.decode("utf-8")
    data = json.loads(json_str)

    assert len(data) == 2
    assert data[0]["tag_no"] == "V-101"
    assert data[1]["tag_no"] == "V-102"
    assert data[0]["qty"] == 2
    assert data[1]["vessel_id_mm"] == 3000.0


def test_export_to_csv():
    """Test exporting to CSV format."""
    records = create_mock_records()
    csv_bytes = export_to_csv(records)

    # Verify we can decode and parse it
    csv_str = csv_bytes.decode("utf-8")
    reader = csv.reader(io.StringIO(csv_str))
    rows = list(reader)

    # Headers + 2 data rows
    assert len(rows) == 3

    # Check headers (assuming order matches definition)
    assert rows[0][0] == "TAG NO."
    assert rows[0][1] == "DESCRIPTION"

    # Check data
    assert rows[1][0] == "V-101"
    assert rows[1][5] == "2"  # qty is cast to string in CSV
    assert rows[2][0] == "V-102"


def test_export_to_excel():
    """Test exporting to Excel format."""
    records = create_mock_records()
    excel_bytes = export_to_excel(records)

    # Verify we can open it with openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(excel_bytes))
    assert "Annexure" in wb.sheetnames

    ws = wb["Annexure"]

    # Check title row
    assert ws.cell(row=1, column=1).value == "BID PROJECT – MECHANICAL DATASHEET ANNEXURE"

    # Check headers (row 2)
    assert ws.cell(row=2, column=1).value == "TAG NO."

    # Check data (row 3 onwards)
    assert ws.cell(row=3, column=1).value == "V-101"

    # Verify numeric types are preserved
    qty_cell = ws.cell(row=3, column=6)  # 6th column is qty
    assert qty_cell.value == 2
    assert isinstance(qty_cell.value, int)

    vessel_id_cell = ws.cell(row=3, column=8)  # 8th column is vessel_id_mm
    assert vessel_id_cell.value == 1500.0
    assert isinstance(vessel_id_cell.value, (int, float))
